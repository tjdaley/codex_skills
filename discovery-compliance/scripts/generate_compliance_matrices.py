from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    raise SystemExit(
        'Missing dependency: ' + str(exc) + '. Install with: '
        'python -m pip install openpyxl'
    )

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
THIN = Side(style='thin', color='D9D9D9')
HEADER_FILL = PatternFill('solid', fgColor='D9EAF7')
TITLE_FILL = PatternFill('solid', fgColor='1F4E78')
TITLE_FONT = Font(color='FFFFFF', bold=True, size=12)
HEADER_FONT = Font(bold=True)


@dataclass
class Statement:
    folder: str
    label: str
    file_path: str
    filename: str
    begin: date | None
    end: date | None
    bates: str
    account_holder: str | None = None


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Build discovery compliance matrices from cleaned CSV exports.')
    parser.add_argument('bank_csv', help='Cleaned bank statement CSV.')
    parser.add_argument('credit_csv', help='Cleaned credit card statement CSV.')
    parser.add_argument('output_dir', help='Folder to receive the matrix outputs.')
    parser.add_argument('--year-start', type=int, default=2020)
    parser.add_argument('--year-end', type=int, default=2026)
    return parser.parse_args()


def parse_date(value: str) -> date:
    return datetime.strptime(value, '%Y-%m-%d').date()


def month_iter(begin: date, end: date):
    year, month = begin.year, begin.month
    while (year, month) <= (end.year, end.month):
        yield year, month
        if month == 12:
            year += 1
            month = 1
        else:
            month += 1


def build_account_label(name: str, account_number: str, fallback: str) -> str:
    clean_name = name.strip()
    clean_account = account_number.strip()
    if clean_name and clean_account:
        return f'{clean_name} x{clean_account}'
    if clean_name:
        return clean_name
    if clean_account:
        return f'x{clean_account}'
    return fallback


def read_statements(csv_path: Path, kind: str) -> list[Statement]:
    with csv_path.open(encoding='utf-8-sig', newline='') as handle:
        rows = list(csv.DictReader(handle))

    statements: list[Statement] = []
    for row in rows:
        folder = Path(row['File Path']).parent.name
        if kind == 'bank':
            label = build_account_label(row['Bank Name'], row['Account Number'], row['Filename Examined'])
        else:
            label = build_account_label(row['Credit Card Issuer'], row['Account Number'], row['Filename Examined'])
        statements.append(
            Statement(
                folder=folder,
                label=label,
                file_path=row['File Path'],
                filename=row['Filename Examined'],
                begin=parse_date(row['Beginning Date of the Statement']) if row['Beginning Date of the Statement'] else None,
                end=parse_date(row['Ending Date of the Statement']) if row['Ending Date of the Statement'] else None,
                bates=row['Bates Number'].strip(),
                account_holder=row['Account Holder(s) Name(s)'].strip() if row['Account Holder(s) Name(s)'].strip() else None
            )
        )
    return statements


def infer_missing_ranges(statements: list[Statement]) -> None:
    grouped: dict[str, list[Statement]] = defaultdict(list)
    for statement in statements:
        grouped[statement.label].append(statement)

    for items in grouped.values():
        items.sort(key=lambda s: ((s.end or date.max), (s.begin or date.max), s.filename))
        prev_end = None
        for item in items:
            if item.begin is None and item.end is not None and prev_end is not None and prev_end < item.end:
                item.begin = prev_end + timedelta(days=1)
            if item.end is not None:
                prev_end = item.end


def split_valid_and_skipped(statements: list[Statement]) -> tuple[list[Statement], list[Statement]]:
    valid: list[Statement] = []
    skipped: list[Statement] = []
    for statement in statements:
        if statement.begin and statement.end and statement.bates:
            valid.append(statement)
        else:
            skipped.append(statement)
    return valid, skipped


def build_matrices(statements: list[Statement], years: list[int]):
    grouped: dict[str, list[Statement]] = defaultdict(list)
    for statement in statements:
        grouped[statement.label].append(statement)

    matrices: dict[str, dict[int, dict[int, list[str]]]] = {}
    for label, items in grouped.items():
        cells = {year: {month: [] for month in range(1, 13)} for year in years}
        seen = {year: {month: set() for month in range(1, 13)} for year in years}
        for item in sorted(items, key=lambda s: (s.begin, s.end, s.bates)):
            for year, month in month_iter(item.begin, item.end):
                if year not in cells:
                    continue
                if item.bates not in seen[year][month]:
                    seen[year][month].add(item.bates)
                    cells[year][month].append(item.bates)
        matrices[label] = cells
    return matrices


def render_markdown(title: str, matrices: dict[str, dict[int, dict[int, list[str]]]], years: list[int]) -> str:
    lines = [f'# {title}', '']
    for label in sorted(matrices):
        lines.append(f'## Account: {label}')
        lines.append('')
        lines.append('Year | ' + ' | '.join(MONTHS))
        lines.append('-----|' + '|'.join(['-----'] * 12))
        for year in years:
            row = [str(year)]
            for month in range(1, 13):
                row.append(', '.join(matrices[label][year][month]))
            lines.append(' | '.join(row))
        lines.append('')
    return '\n'.join(lines).rstrip() + '\n'


def style_range(ws, row_idx, start_col, end_col, fill=None, font=None, align=None):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if align:
            cell.alignment = align
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def add_sheet(ws, title: str, matrices: dict[str, dict[int, dict[int, list[str]]]], years: list[int]):
    ws.title = title
    row = 1
    for label in sorted(matrices):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        ws.cell(row=row, column=1, value=f'Account: {label}')
        style_range(ws, row, 1, 13, fill=TITLE_FILL, font=TITLE_FONT, align=Alignment(horizontal='left'))
        row += 1

        headers = ['Year'] + MONTHS
        for col, value in enumerate(headers, start=1):
            ws.cell(row=row, column=col, value=value)
        style_range(ws, row, 1, 13, fill=HEADER_FILL, font=HEADER_FONT, align=Alignment(horizontal='center'))
        row += 1

        for year in years:
            ws.cell(row=row, column=1, value=str(year))
            ws.cell(row=row, column=1).border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            for month in range(1, 13):
                cell = ws.cell(row=row, column=month + 1, value=', '.join(matrices[label][year][month]))
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            row += 1
        row += 2

    widths = [12] + [18] * 12
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    ws.freeze_panes = 'A2'


def write_excel(out_path: Path, bank_matrices, credit_matrices, years: list[int]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_sheet(wb.create_sheet(), 'Bank Accounts', bank_matrices, years)
    add_sheet(wb.create_sheet(), 'Credit Cards', credit_matrices, years)
    wb.save(out_path)


def main() -> int:
    args = parse_args()
    years = list(range(args.year_start, args.year_end + 1))
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    bank_raw = read_statements(Path(args.bank_csv), 'bank')
    credit_raw = read_statements(Path(args.credit_csv), 'credit')
    infer_missing_ranges(bank_raw)
    infer_missing_ranges(credit_raw)
    bank_valid, bank_skipped = split_valid_and_skipped(bank_raw)
    credit_valid, credit_skipped = split_valid_and_skipped(credit_raw)

    bank_matrices = build_matrices(bank_valid, years)
    credit_matrices = build_matrices(credit_valid, years)

    bank_md = output_dir / 'bank_compliance_matrices.md'
    credit_md = output_dir / 'credit_card_compliance_matrices.md'
    xlsx_path = output_dir / 'compliance_matrices.xlsx'

    bank_md.write_text(render_markdown('Bank Account Compliance Matrices', bank_matrices, years), encoding='utf-8')
    credit_md.write_text(render_markdown('Credit Card Compliance Matrices', credit_matrices, years), encoding='utf-8')
    write_excel(xlsx_path, bank_matrices, credit_matrices, years)

    print(f'Bank markdown: {bank_md}')
    print(f'Credit markdown: {credit_md}')
    print(f'Excel workbook: {xlsx_path}')
    print(f'Bank statements used={len(bank_valid)} skipped={len(bank_skipped)}')
    print(f'Credit card statements used={len(credit_valid)} skipped={len(credit_skipped)}')
    for statement in bank_skipped:
        print('Skipped bank:', statement.filename)
    for statement in credit_skipped:
        print('Skipped credit:', statement.filename)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
